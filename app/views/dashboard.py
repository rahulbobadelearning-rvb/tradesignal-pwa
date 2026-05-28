# REMARK: The dashboard is the executive overview.
# It adds: Excel export of filtered data, and one-click PO risk email alerts.

import html as _html
import io
from datetime import date, datetime
from typing import Optional

import pandas as pd
import streamlit as st

from core.analytics import get_dashboard_rows
from core.db import get_connection
from core.email_utils import send_risk_alert, smtp_configured
from core.ui_tokens import (
    COLOR_DANGER,
    COLOR_SUCCESS,
    COLOR_TEXT_SECONDARY,
    COLOR_WARNING,
    status_badge,
)


def render() -> None:
    st.markdown(
        "<h1 style='font-size:1.6rem; font-weight:800; margin-bottom:4px;'>Dashboard</h1>"
        f"<p style='color:{COLOR_TEXT_SECONDARY}; font-size:0.85rem; margin-top:0;'>"
        "Vendor PO health · spend runway · YTD consumption</p>",
        unsafe_allow_html=True,
    )

    rows = get_dashboard_rows()

    if not rows:
        st.info("No vendors yet. Add vendors in **Vendor Master** to see analytics here.")
        return

    _render_kpis(rows)
    st.markdown("<hr class='mds-divider'>", unsafe_allow_html=True)

    # ── Filter bar ────────────────────────────────────────────────────────────
    col_filter, col_sort, col_export = st.columns([3, 2, 1])
    with col_filter:
        search = st.text_input("Filter vendors",
                                placeholder="Vendor name, code, or country…",
                                label_visibility="collapsed")
    with col_sort:
        sort_by = st.selectbox(
            "Sort by",
            ["Vendor Name", "YTD Spend", "Remaining PO",
             "Runway Status", "PO Expiration", "Country"],
            label_visibility="collapsed",
        )

    df = pd.DataFrame(rows)

    if search:
        mask = (
            df["Vendor Name"].str.contains(search, case=False, na=False)
            | df["Vendor Code"].str.contains(search, case=False, na=False)
            | df.get("Country", pd.Series([""] * len(df))).str.contains(
                search, case=False, na=False)
        )
        df = df[mask]

    ascending = sort_by not in ("YTD Spend",)
    df = df.sort_values(sort_by, ascending=ascending, ignore_index=True)

    if df.empty:
        st.warning("No vendors match the filter.")
        return

    # ── Excel export button ───────────────────────────────────────────────────
    with col_export:
        excel_bytes = _build_excel(df)
        st.download_button(
            label="⬇️ Excel",
            data=excel_bytes,
            file_name=f"vendor_dashboard_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Export currently displayed vendors to Excel",
        )

    # ── Vendor table ──────────────────────────────────────────────────────────
    st.markdown("<div class='mds-section-title'>Vendor Spend Overview</div>",
                unsafe_allow_html=True)
    _render_vendor_table(df)

    # ── Risk panel + email alerts ─────────────────────────────────────────────
    _render_risk_panel(df, rows)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _build_excel(df: pd.DataFrame) -> bytes:
    """Build a formatted Excel workbook from the current dashboard view."""
    export_df = pd.DataFrame({
        "Vendor Name":      df["Vendor Name"],
        "Vendor Code":      df["Vendor Code"],
        "Country":          df.get("Country", "—"),
        "PO Number":        df["PO Number"],
        "PO Value ($)":     df["PO Value"].apply(lambda x: round(x, 2)),
        "PO Expiration":    df["PO Expiration"],
        "Avg Monthly ($)":  df["Avg Monthly"].apply(lambda x: round(x, 2)),
        "Last Monthly ($)": df["Last Monthly"].apply(lambda x: round(x, 2)),
        "YTD Spend ($)":    df["YTD Spend"].apply(lambda x: round(x, 2)),
        "Remaining PO ($)": df["Remaining PO"].apply(lambda x: round(x, 2)),
        "Months Left":      df["Months Left"],
        "Expected Monthly ($)": df["Expected Monthly"].apply(
            lambda x: round(x, 2) if x is not None else ""),
        "Application Owner": df.get("Application Owner", ""),
        "Runway Status":    df["Runway Status"],
    })

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Dashboard")
        ws = writer.sheets["Dashboard"]

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col_cells if cell.value is not None),
                default=10,
            )
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

        # Bold header row
        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="0073AB")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Colour-code Runway Status column (last column)
        status_col = export_df.columns.get_loc("Runway Status") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=status_col)
            if cell.value == "Risk":
                cell.font = Font(bold=True, color="B45309")
                cell.fill = PatternFill("solid", fgColor="FEF3C7")
            elif cell.value == "Expired":
                cell.font = Font(bold=True, color="B91C1C")
                cell.fill = PatternFill("solid", fgColor="FEE2E2")
            elif cell.value == "On Track":
                cell.font = Font(bold=True, color="1D6F42")
                cell.fill = PatternFill("solid", fgColor="D1FAE5")

    return buf.getvalue()


# ---------------------------------------------------------------------------
# KPI strip
# ---------------------------------------------------------------------------

def _render_kpis(rows: list[dict]) -> None:
    total_vendors = len(rows)
    total_ytd = sum(r["YTD Spend"] for r in rows)
    total_po = sum(r["PO Value"] for r in rows)
    at_risk = sum(1 for r in rows if r["Runway Status"] in ("Risk", "Expired"))

    with get_connection() as conn:
        invoice_count = conn.execute("SELECT COUNT(*) FROM invoices").fetchone()[0]

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("Total Vendors", total_vendors)
    c2.metric("Total PO Value", f"${total_po:,.0f}")
    c3.metric("YTD Spend", f"${total_ytd:,.0f}")
    c4.metric("Invoices Uploaded", invoice_count)
    c5.metric("At-Risk Vendors", at_risk)


# ---------------------------------------------------------------------------
# Vendor table
# ---------------------------------------------------------------------------

def _render_vendor_table(df: pd.DataFrame) -> None:
    display = pd.DataFrame()
    display["Vendor Name"]   = df["Vendor Name"]
    display["Vendor Code"]   = df["Vendor Code"]
    display["Country"]       = df.get("Country", "—")
    display["PO Value"]      = df["PO Value"].apply(lambda x: f"${x:,.0f}")
    display["PO Expiration"] = df["PO Expiration"]
    display["Avg Monthly"]   = df["Avg Monthly"].apply(lambda x: f"${x:,.0f}")
    display["Last Monthly"]  = df["Last Monthly"].apply(lambda x: f"${x:,.0f}")
    display["YTD Spend"]     = df["YTD Spend"].apply(lambda x: f"${x:,.0f}")
    display["Remaining PO"]  = df["Remaining PO"].apply(
        lambda x: f"${x:,.0f}" if x >= 0 else f"-${abs(x):,.0f}"
    )
    display["Months Left"]   = df["Months Left"].apply(
        lambda x: str(x) if x is not None else "—"
    )
    display["Runway"]        = df["Runway Status"]

    def colour_runway(val: str) -> str:
        if val == "On Track":  return f"color:{COLOR_SUCCESS};font-weight:700;"
        if val == "Risk":      return f"color:{COLOR_WARNING};font-weight:700;"
        if val == "Expired":   return f"color:{COLOR_DANGER};font-weight:700;"
        return ""

    def highlight_remaining(val: str) -> str:
        if str(val).startswith("-"):
            return f"background-color:#FEE2E2;color:{COLOR_DANGER};font-weight:700;"
        return ""

    styled = (
        display.style
        .map(colour_runway, subset=["Runway"])
        .map(highlight_remaining, subset=["Remaining PO"])
    )
    st.dataframe(styled, use_container_width=True, hide_index=True, height=400)


# ---------------------------------------------------------------------------
# Risk panel + email alerts
# ---------------------------------------------------------------------------

def _render_risk_panel(df: pd.DataFrame, all_rows: list[dict]) -> None:
    at_risk = df[df["Runway Status"].isin(["Risk", "Expired"])]
    expiring_soon = df[df["PO Expiration"].apply(_expires_within_60_days)]

    if at_risk.empty and expiring_soon.empty:
        st.success("✅ All vendors are on track. No immediate risks detected.")
        return

    st.markdown("<hr class='mds-divider'>", unsafe_allow_html=True)
    st.markdown("<div class='mds-section-title'>Risk Alerts</div>",
                unsafe_allow_html=True)

    c1, c2 = st.columns(2)

    with c1:
        if not at_risk.empty:
            st.markdown(
                f"<div style='color:{COLOR_DANGER}; font-weight:700; margin-bottom:8px;'>"
                f"⚠ Budget Risk ({len(at_risk)} vendor{'s' if len(at_risk)>1 else ''})</div>",
                unsafe_allow_html=True,
            )
            for _, row in at_risk.iterrows():
                safe_name = _html.escape(str(row["Vendor Name"]))
                safe_status = _html.escape(str(row["Runway Status"]))
                safe_remaining = _html.escape(str(row["Remaining PO"]))
                st.markdown(
                    f"**{safe_name}** — {safe_status} · Remaining: {safe_remaining}"
                )

            # ── Email alert buttons ───────────────────────────────────────
            st.markdown("<hr class='mds-divider'>", unsafe_allow_html=True)
            st.markdown("**Send PO Risk Alert Email**")

            if not smtp_configured():
                st.caption("⚙️ Configure SMTP in **Settings** to enable email alerts.")
            else:
                # Build vendor lookup by name
                vendor_id_map = {r["Vendor Name"]: r["id"] for r in all_rows}

                for _, row in at_risk.iterrows():
                    vendor_name = row["Vendor Name"]
                    vid = vendor_id_map.get(vendor_name)
                    btn_key = f"email_{vid}"

                    col_btn, col_status = st.columns([2, 3])
                    with col_btn:
                        if st.button(
                            f"📧 Alert: {vendor_name[:22]}",
                            key=btn_key,
                            use_container_width=True,
                        ):
                            with st.spinner(f"Sending alert for {vendor_name}…"):
                                error = send_risk_alert(vid)
                            if error:
                                st.session_state[f"email_result_{vid}"] = ("error", error)
                            else:
                                st.session_state[f"email_result_{vid}"] = ("ok", "")

                    with col_status:
                        result = st.session_state.get(f"email_result_{vid}")
                        if result:
                            status, msg = result
                            if status == "ok":
                                st.success("✅ Alert sent to Rahul.bobade@maersk.com")
                            else:
                                st.error(f"❌ Send failed — check SMTP in Settings. ({msg[:80]})")

    with c2:
        if not expiring_soon.empty:
            st.markdown(
                f"<div style='color:{COLOR_WARNING}; font-weight:700; margin-bottom:8px;'>"
                f"⏰ PO Expiring Soon ({len(expiring_soon)} vendor{'s' if len(expiring_soon)>1 else ''})</div>",
                unsafe_allow_html=True,
            )
            for _, row in expiring_soon.iterrows():
                safe_name = _html.escape(str(row["Vendor Name"]))
                safe_expiry = _html.escape(str(row["PO Expiration"]))
                st.markdown(f"**{safe_name}** — expires {safe_expiry}")


def _expires_within_60_days(expiry_str: str) -> bool:
    if expiry_str in ("—", "", None):
        return False
    try:
        expiry = datetime.strptime(expiry_str, "%Y-%m-%d").date()
        delta = (expiry - date.today()).days
        return 0 <= delta <= 60
    except ValueError:
        return False
